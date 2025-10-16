from openpyxl import load_workbook
import xlrd
import os
import unicodedata
from typing import List, Tuple
from datetime import datetime, timedelta
from app.schemas.upload import BulkDataSchema, ErrorDetail
import logging

logger = logging.getLogger(__name__)

# Mapeo limpio de encabezados del archivo a campos internos
HEADER_MAP = {
    "id": "codigo_cliente",
    "nombre_completo": "nombre_completo",
    "fecha_de_nacimiento": "fecha_nacimiento",
    "direccion": "direccion",
    "localidad_y_codigo_postal": "localidad_cp",
    "telefono": "telefono",
    "correo_electronico": "email",
    "fecha_de_alta": "fecha_alta",
    "grupo_de_clientes": "grupo_clientes",
    "valor": "valor",
    "descripcion": "descripcion"
}

REQUIRED_FIELDS = ["nombre_completo", "email"]

def normalize_header(header: str) -> str:
    """Normaliza encabezados eliminando tildes, espacios y convirtiendo a minúsculas"""
    header = unicodedata.normalize("NFKD", header).encode("ASCII", "ignore").decode("utf-8")
    return header.lower().strip().replace(" ", "_")

def excel_date_to_date(value: float) -> datetime.date:
    """Convierte fecha serial de Excel a objeto date"""
    if isinstance(value, float):
        base_date = datetime(1899, 12, 30)
        return (base_date + timedelta(days=value)).date()
    return value

class FileProcessor:
    """Servicio para validar y extraer datos de archivos XLS/XLSX"""

    @staticmethod
    def validate_file(file_path: str) -> bool:
        """Valida que el archivo sea XLS o XLSX válido"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in [".xlsx", ".xls"]:
            raise ValueError("El archivo debe ser .xlsx o .xls")

        try:
            if ext == ".xlsx":
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                if not workbook.sheetnames:
                    raise ValueError("El archivo no contiene hojas")
            else:
                workbook = xlrd.open_workbook(file_path)
                if workbook.nsheets == 0:
                    raise ValueError("El archivo no contiene hojas")
            return True
        except Exception as e:
            logger.error(f"Error validando archivo: {e}")
            raise ValueError(f"Archivo inválido: {e}")

    @staticmethod
    def extract_data(file_path: str, batch_id: str) -> Tuple[List[dict], List[ErrorDetail]]:
        """Extrae y valida datos del archivo XLS/XLSX"""
        valid_data = []
        errors = []

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            raw_headers = [str(cell.value).strip() for cell in worksheet[1] if cell.value]
            headers = [HEADER_MAP.get(normalize_header(h), normalize_header(h)) for h in raw_headers]
            rows = worksheet.iter_rows(min_row=2, values_only=True)
        else:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            raw_headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
            headers = [HEADER_MAP.get(normalize_header(h), normalize_header(h)) for h in raw_headers]
            rows = (sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows))

        for row_idx, row in enumerate(rows, start=2):
            try:
                raw_data = {headers[i]: row[i] for i in range(len(headers))}
                raw_data["batch_id"] = batch_id

                # Normalizar fechas
                for field in ["fecha_nacimiento", "fecha_alta"]:
                    if field in raw_data and isinstance(raw_data[field], float):
                        raw_data[field] = excel_date_to_date(raw_data[field])

                # Validar campos obligatorios
                for field in REQUIRED_FIELDS:
                    if not raw_data.get(field):
                        raise ValueError(f"Campo obligatorio '{field}' vacío")

                # Validar con Pydantic
                record = BulkDataSchema(**raw_data)
                valid_data.append(record.dict())

            except Exception as e:
                errors.append(ErrorDetail(
                    row_number=row_idx,
                    error=str(e),
                    data=dict(zip(headers, row))
                ))

        logger.info(f"Procesamiento completado: {len(valid_data)} válidos, {len(errors)} errores")
        return valid_data, errors
